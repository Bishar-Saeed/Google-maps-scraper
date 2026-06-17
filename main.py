#!/usr/bin/env python3
"""
OPTIMIZED BUT THOROUGH EMAIL SCRAPER
- Checks 5-6 website pages (auto-discovery + fallback)
- ALWAYS checks Facebook
- Optimized but doesn't skip any data
- Balanced speed: not too fast (avoids blocks), not too slow (efficient)
- IMPROVED SCROLLING: Gets ALL Google Maps results (up to 120 per search)
"""

import asyncio
import re
import os
import random
from datetime import datetime
from typing import List, Dict, Optional, Set, Tuple
from urllib.parse import urlparse, urljoin
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================================
# CONFIGURATION - OPTIMIZED BUT THOROUGH
# ============================================================================

# DON'T SKIP ANYTHING - Scrape all listings
SKIP_LISTINGS_CONTAINING = []  # EMPTY - Scrape everything!

# Only skip obvious junk emails (won't skip real emails)
SKIP_EMAILS_CONTAINING = [
    "sentry", "mailer-daemon", "postmaster", "noreply", "no-reply",
    "donotreply", "example.com", "test.com", "placeholder"
]

# Skip only obvious spam domains
SKIP_EMAIL_DOMAINS = [
    "wixpress.com", "sentry.io", "example.com", "example.org"
]

# OPTIMIZED TIMING - Fast enough but won't get blocked
BETWEEN_LISTINGS = random.uniform(5.0, 8.0)  # 5-8 seconds (optimized)
BETWEEN_PAGES = random.uniform(1.0, 1.5)     # 1-1.5 seconds between pages
NAVIGATION_TIMEOUT = 45000                   # 45 seconds (optimized)
PAGE_LOAD_TIMEOUT = 20000                    # 20 seconds per page
MAX_RETRIES = 2                              # Retry failed pages twice
MAX_PAGES_TO_CHECK = 6                       # Always check 6 pages

# SCROLLING SETTINGS - Gets ALL results
SCROLL_DELAY = random.uniform(2.0, 2.8)      # Medium pace scrolling (2-2.8 seconds)
MAX_SCROLL_ATTEMPTS = 30                      # Enough for ~120 results
NO_NEW_RESULTS_LIMIT = 4                      # Stop after 4 scrolls with no new results

# Parallel settings (for concurrent page checks - FASTER but safe)
MAX_CONCURRENT_PAGES = 2  # Check 2 pages at a time max (safe concurrency)

# US States Database
US_STATES = {
    "California": ["Los Angeles",],
    "New York": ["New York", "Buffalo", "Rochester"],
    "Texas": ["Houston", "San Antonio", "Dallas", "Austin"],
    "Florida": ["Miami", "Orlando", "Tampa", "Jacksonville"],
    "Illinois": ["Chicago", "Aurora", "Rockford", "Joliet", "Naperville", "Springfield", "Peoria", "Elgin"],
}

# Google Maps Selectors
SELECTORS = {
    "results_panel": 'div[role="feed"]',
    "result_link": 'a[href*="/maps/place/"]',
    "name": 'h1.DUwDvf, h1[aria-level="1"]',
    "rating": 'div.F7nice span[aria-hidden="true"]',
    "reviews": 'button[aria-label*="reviews"] span',
    "address": 'button[data-item-id="address"] div.Io6YTe',
    "phone": 'button[data-item-id*="phone"] div.Io6YTe',
    "website": 'a[data-item-id="authority"]',
    "category": 'button[jsaction*="category"] span',
}

# Excel columns
COLUMNS = [
    "No.", "Listing Name", "Category", "Rating", "Reviews", "Address",
    "Phone", "Website", "All Emails Found", "Website Emails",
    "Facebook Emails", "Primary Email", "City", "State", "Google Maps URL", "Scrape Date"
]

# FALLBACK PAGES - Used ONLY when auto-discovery fails
FALLBACK_PAGES = [
    "", "contact", "contact-us", "about", "about-us", "team", "our-team",
    "meet-the-team", "staff", "doctors", "providers", "location", "hours"
]

# ============================================================================
# OPTIMIZED PAGE DISCOVERY (Fast but thorough)
# ============================================================================

async def discover_pages_optimized(page, base_url: str, max_pages: int = None) -> List[str]:
    """
    OPTIMIZED page discovery:
    - Quick homepage scan (3 second timeout)
    - Parallel link extraction
    - Immediate fallback if needed
    """
    discovered_urls = set()
    discovered_urls.add(base_url)
    
    try:
        # Fast homepage load (reduced timeout for speed)
        await page.goto(base_url, timeout=15000, wait_until="domcontentloaded")
        await asyncio.sleep(random.uniform(0.5, 0.8))  # Minimal pause
        
        # Quick link extraction (optimized)
        links = await page.evaluate('''
            () => {
                const links = new Set();
                const baseUrl = window.location.origin;
                const allLinks = document.querySelectorAll('a[href]');
                const skipExtensions = ['.jpg', '.png', '.gif', '.pdf', '.css', '.js', '.xml'];
                
                for (let link of allLinks) {
                    let href = link.href;
                    if (href && href.startsWith(baseUrl) && 
                        !href.includes('#') && 
                        !href.includes('?') &&
                        href !== window.location.href) {
                        
                        href = href.replace(/\\/$/, '');
                        const isContent = !skipExtensions.some(ext => href.toLowerCase().endsWith(ext));
                        if (isContent) links.add(href);
                    }
                    if (links.size >= 20) break;  // Limit links for speed
                }
                return Array.from(links);
            }
        ''')
        
        # Add discovered links (up to max_pages)
        for link in links[:max_pages - 1]:
            if link and link != base_url:
                discovered_urls.add(link)
                
        if len(discovered_urls) > 1:
            print(f"      ✅ Found {len(discovered_urls) - 1} internal pages")
        else:
            print(f"      ⚠️ No internal links found, using fallback")
            
    except Exception as e:
        print(f"      ⚠️ Auto-discovery failed, using fallback: {str(e)[:40]}")
    
    # Add fallback pages if needed
    if len(discovered_urls) < max_pages:
        needed = max_pages - len(discovered_urls)
        for page_path in FALLBACK_PAGES[:needed]:
            if page_path:
                full_url = urljoin(base_url, page_path)
                if full_url not in discovered_urls:
                    discovered_urls.add(full_url)
    
    # Convert to list and return
    result = list(discovered_urls)[:max_pages]
    
    # Brief summary (minimal printing for speed)
    print(f"      📄 Checking {len(result)} pages", end="")
    if len(discovered_urls) > 1:
        print(f" ({len(discovered_urls) - 1} discovered)")
    else:
        print(f" (using fallback)")
    
    return result

# ============================================================================
# OPTIMIZED WEBSITE EMAIL EXTRACTION
# ============================================================================

async def extract_emails_from_website_optimized(page, website_url: str, listing_name: str) -> Tuple[List[str], List[str]]:
    """
    OPTIMIZED email extraction:
    - Quick page discovery
    - Sequential page checking (safer, avoids detection)
    - No data skipped
    """
    if not website_url:
        return [], []
    
    base_url = website_url.rstrip('/')
    email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
    
    # Discover pages quickly
    pages_to_check = await discover_pages_optimized(page, base_url, MAX_PAGES_TO_CHECK)
    
    # Check pages sequentially (safer, avoids detection)
    all_emails = set()
    successful_pages = 0
    
    for i, page_url in enumerate(pages_to_check, 1):
        try:
            # Navigate to page
            await page.goto(page_url, timeout=PAGE_LOAD_TIMEOUT, wait_until="domcontentloaded")
            await asyncio.sleep(random.uniform(0.3, 0.5))
            
            # Get content
            content = await page.content()
            emails = re.findall(email_pattern, content)
            
            for email in emails:
                email_lower = email.lower()
                
                # Skip junk
                if any(x in email_lower for x in SKIP_EMAILS_CONTAINING):
                    continue
                
                domain = email.split('@')[-1] if '@' in email else ""
                if any(skip in domain for skip in SKIP_EMAIL_DOMAINS):
                    continue
                
                all_emails.add(email)
            
            successful_pages += 1
            
            # Progress indicator (minimal)
            if i % 3 == 0:
                print(f"      Checked {i}/{len(pages_to_check)} pages...", end="\r")
            
        except Exception:
            continue
    
    print(f"      ✅ Checked {successful_pages}/{len(pages_to_check)} pages    ")
    
    email_list = list(all_emails)
    
    if email_list:
        print(f"   📧 Found {len(email_list)} email(s) on website")
        for e in email_list[:3]:
            print(f"      - {e}")
    
    return email_list, email_list

# ============================================================================
# OPTIMIZED FACEBOOK CHECK (Always checked)
# ============================================================================

async def extract_emails_from_facebook_optimized(page, listing_name: str, website_url: str) -> List[str]:
    """
    OPTIMIZED Facebook check:
    - Single attempt per URL pattern
    - Quick timeout (10 seconds)
    - Still checks thoroughly
    """
    emails_found = set()
    email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
    
    # Clean listing name
    name_clean = re.sub(r'[^\w\s]', '', listing_name.lower())
    name_clean = re.sub(r'\s+', '', name_clean)
    
    # Extract domain
    domain = ""
    if website_url:
        try:
            parsed = urlparse(website_url)
            domain = parsed.netloc.replace('www.', '').split('.')[0]
        except:
            pass
    
    # Facebook URLs to try (prioritized)
    facebook_urls = [
        f"https://www.facebook.com/{name_clean}",
        f"https://www.facebook.com/{domain}",
    ]
    
    print(f"   🔍 Checking Facebook...")
    
    for fb_url in facebook_urls[:2]:  # Only 2 attempts max
        try:
            await page.goto(fb_url, timeout=10000, wait_until="domcontentloaded")
            await asyncio.sleep(random.uniform(0.5, 0.8))
            
            content = await page.content()
            emails = re.findall(email_pattern, content)
            
            for email in emails:
                email_lower = email.lower()
                if not any(x in email_lower for x in SKIP_EMAILS_CONTAINING):
                    if 'facebook.com' not in email_lower and 'fbcdn' not in email_lower:
                        emails_found.add(email)
            
        except Exception:
            continue
    
    email_list = list(emails_found)
    
    if email_list:
        print(f"   📧 Found {len(email_list)} email(s) on Facebook")
        for e in email_list[:2]:
            print(f"      - {e}")
    
    return email_list

# ============================================================================
# COMPLETE EMAIL EXTRACTION
# ============================================================================

async def extract_all_emails_optimized(page, website_url: str, listing_name: str) -> Tuple[List[str], List[str], List[str]]:
    """Extract all emails using optimized methods."""
    
    all_emails_set = set()
    
    # Extract from website
    website_emails, _ = await extract_emails_from_website_optimized(page, website_url, listing_name)
    all_emails_set.update(website_emails)
    
    # Brief pause before Facebook
    await asyncio.sleep(random.uniform(0.5, 0.8))
    
    # Extract from Facebook
    facebook_emails = await extract_emails_from_facebook_optimized(page, listing_name, website_url)
    all_emails_set.update(facebook_emails)
    
    all_emails_list = list(all_emails_set)
    
    # Determine primary email (quick)
    primary_email = ""
    if all_emails_list:
        for email in all_emails_list:
            if '.com' in email and not any(x in email.lower() for x in ["info", "admin", "contact"]):
                primary_email = email
                break
        if not primary_email:
            primary_email = all_emails_list[0]
    
    print(f"\n   📊 EMAIL SUMMARY: {len(all_emails_list)} total (web: {len(website_emails)}, FB: {len(facebook_emails)})")
    
    return all_emails_list, website_emails, facebook_emails

# ============================================================================
# SCRAPE LISTING (Optimized)
# ============================================================================

async def scrape_listing_optimized(page, url: str, city: str, state: str, index: int, total: int,
                                    wb, ws, filename: str, retry_count: int = 0) -> Optional[Dict]:
    """Scrape a single listing with optimized settings."""
    
    print(f"\n[{index}/{total}] 📍 {city[:15]} - {url.split('/')[-2][:35]}...")
    
    try:
        # Fast Google Maps load
        await page.goto(url, timeout=NAVIGATION_TIMEOUT, wait_until="domcontentloaded")
        await asyncio.sleep(random.uniform(1.0, 1.5))
        
        # Extract basic info (quick)
        listing_name = await get_text_fast(page, SELECTORS["name"])
        if not listing_name:
            print(f"   ❌ No name found")
            return None
        
        print(f"   🏢 {listing_name[:55]}")
        
        # Extract other details
        category = await get_text_fast(page, SELECTORS["category"])
        rating = await get_text_fast(page, SELECTORS["rating"])
        reviews_raw = await get_text_fast(page, SELECTORS["reviews"])
        reviews = re.sub(r'[^\d]', '', reviews_raw) if reviews_raw else ""
        address = await get_text_fast(page, SELECTORS["address"])
        phone = await get_text_fast(page, SELECTORS["phone"])
        
        # Extract website
        website = await extract_website_fast(page)
        
        if not website:
            print(f"   ⏭️ No website found")
            return None
        
        print(f"   🌐 {website[:60]}...")
        
        # Create new page for email extraction
        email_page = await page.context.new_page()
        
        # Extract ALL emails (optimized)
        all_emails, website_emails, facebook_emails = await extract_all_emails_optimized(
            email_page, website, listing_name
        )
        
        await email_page.close()
        
        if not all_emails:
            print(f"   ❌ No emails found")
            return None
        
        # Prepare data
        data = {
            "listing_name": listing_name,
            "category": category,
            "rating": rating,
            "reviews": reviews,
            "address": address,
            "phone": phone,
            "website": website,
            "all_emails": ", ".join(all_emails),
            "website_emails": ", ".join(website_emails),
            "facebook_emails": ", ".join(facebook_emails),
            "primary_email": all_emails[0] if all_emails else "",
            "city": city,
            "state": state,
            "url": url,
            "scrape_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
        print(f"\n   ✅ FOUND {len(all_emails)} emails! (Saved)")
        
        # Save to Excel
        save_to_excel_fast(wb, ws, data, filename)
        
        return data
        
    except Exception as e:
        print(f"   ❌ Error: {str(e)[:60]}")
        if retry_count < MAX_RETRIES:
            print(f"   🔄 Retry {retry_count + 1}/{MAX_RETRIES}...")
            await asyncio.sleep(3)
            return await scrape_listing_optimized(page, url, city, state, index, total, wb, ws, filename, retry_count + 1)
        return None

# ============================================================================
# FAST HELPER FUNCTIONS (Optimized)
# ============================================================================

async def extract_website_fast(page) -> str:
    """Extract website URL quickly."""
    selectors = ['a[data-item-id="authority"]', 'a[aria-label*="Website"]']
    
    for selector in selectors:
        try:
            element = await page.query_selector(selector)
            if element:
                href = await element.get_attribute("href")
                if href and ("http" in href or "www" in href):
                    if href.startswith("//"):
                        href = "https:" + href
                    return href.rstrip('/')
        except:
            continue
    return ""

async def get_text_fast(page, *selectors) -> str:
    """Get text quickly with fallback."""
    for selector in selectors:
        try:
            element = await page.query_selector(selector)
            if element:
                text = await element.inner_text()
                if text and text.strip():
                    return text.strip()[:500]
        except:
            continue
    return ""

saved_records_count = 0

def save_to_excel_fast(wb, ws, data: Dict, filename: str):
    """Save data to Excel quickly."""
    global saved_records_count
    
    next_row = ws.max_row + 1
    
    values = [
        next_row - 1,
        data.get("listing_name", ""),
        data.get("category", ""),
        data.get("rating", ""),
        data.get("reviews", ""),
        data.get("address", ""),
        data.get("phone", ""),
        data.get("website", ""),
        data.get("all_emails", ""),
        data.get("website_emails", ""),
        data.get("facebook_emails", ""),
        data.get("primary_email", ""),
        data.get("city", ""),
        data.get("state", ""),
        data.get("url", ""),
        data.get("scrape_date", ""),
    ]
    
    for col_idx, value in enumerate(values, 1):
        ws.cell(row=next_row, column=col_idx, value=value)
    
    try:
        wb.save(filename)
        saved_records_count += 1
    except Exception as e:
        print(f"   ⚠️ Save error: {e}")

# ============================================================================
# IMPROVED URL COLLECTION WITH PROPER SCROLLING - GETS ALL RESULTS
# ============================================================================

async def collect_urls_optimized(page, profession: str, state: str, city: str, limit: int = 120) -> List[str]:
    """
    IMPROVED URL collection with MEDIUM-PACED scrolling to get ALL results.
    Google Maps shows ~120 results maximum per search.
    This function scrolls properly to load ALL of them.
    """
    print(f"\n   🌆 Searching: {profession} in {city}, {state}")
    
    query = f"{profession} in {city}, {state}"
    search_url = f"https://www.google.com/maps/search/{query.replace(' ', '+')}/"
    
    try:
        await page.goto(search_url, timeout=30000, wait_until="domcontentloaded")
        await asyncio.sleep(random.uniform(2.5, 3.5))  # Longer initial wait
    except:
        print(f"   ⚠️ Failed to load search")
        return []
    
    # Handle cookie consent if present
    try:
        accept_btn = await page.query_selector('button:has-text("Accept all")')
        if accept_btn:
            await accept_btn.click()
            await asyncio.sleep(1)
    except:
        pass
    
    # Get the results panel (important - scroll this, not the whole page)
    results_panel = await page.query_selector('div[role="feed"]')
    if not results_panel:
        print(f"   ⚠️ Could not find results panel, trying page scroll...")
        results_panel = None
    
    urls = set()
    previous_count = 0
    no_new_count = 0
    scroll_attempts = 0
    
    print(f"      📜 Starting medium-pace scrolling to collect ALL results...")
    print(f"      ⏱️  {SCROLL_DELAY:.1f}s between scrolls (avoids detection)")
    
    while len(urls) < limit and scroll_attempts < MAX_SCROLL_ATTEMPTS and no_new_count < NO_NEW_RESULTS_LIMIT:
        
        # Extract current URLs
        links = await page.query_selector_all('a[href*="/maps/place/"]')
        
        for link in links:
            href = await link.get_attribute("href")
            if href and "/maps/place/" in href:
                clean_url = href.split("?")[0]
                urls.add(clean_url)
        
        # Show progress
        print(f"      📊 Collected {len(urls)} URLs...", end="\r")
        
        # Check if we got new results
        if len(urls) == previous_count:
            no_new_count += 1
            if no_new_count == 1:
                print(f"\n      ⏸️ No new results after scroll {scroll_attempts + 1}")
        else:
            new_found = len(urls) - previous_count
            print(f"\n      ✅ +{new_found} new results (Total: {len(urls)})")
            no_new_count = 0
        
        previous_count = len(urls)
        
        # Stop if we have enough or reached end
        if len(urls) >= limit:
            break
        
        if no_new_count >= NO_NEW_RESULTS_LIMIT:
            print(f"\n      🏁 Reached end of results (no new listings after {NO_NEW_RESULTS_LIMIT} scrolls)")
            break
        
        # SCROLLING - Medium pace
        if results_panel:
            # Scroll the results panel (more reliable)
            try:
                await results_panel.evaluate("""
                    (element) => {
                        const previousHeight = element.scrollHeight;
                        element.scrollTop = element.scrollHeight;
                        return previousHeight;
                    }
                """)
            except:
                # Fallback to page scroll
                await page.evaluate("window.scrollBy(0, 800)")
        else:
            # Scroll the whole page
            await page.evaluate("window.scrollBy(0, 800)")
        
        # Medium pace delay - allows Google to load new results
        await asyncio.sleep(SCROLL_DELAY)
        
        scroll_attempts += 1
        
        # Progress update every 5 scrolls
        if scroll_attempts % 5 == 0:
            print(f"      📜 Scroll {scroll_attempts}/{MAX_SCROLL_ATTEMPTS} - {len(urls)} URLs so far")
    
    print(f"\n      ✅ Scrolling complete! Final count: {len(urls)} URLs")
    
    # Show statistics
    if len(urls) >= 100:
        print(f"      🎯 Excellent! Got {len(urls)} results (near maximum of ~120)")
    elif len(urls) >= 60:
        print(f"      👍 Good! Got {len(urls)} results")
    elif len(urls) >= 30:
        print(f"      📊 Got {len(urls)} results - moderate coverage")
    else:
        print(f"      ⚠️ Got only {len(urls)} results - try a broader search term")
    
    return list(urls)

# ============================================================================
# URL COLLECTION FROM ALL CITIES
# ============================================================================

async def collect_all_urls_optimized(page, profession: str, state: str, cities: List[str], target: int) -> List[str]:
    """Collect URLs from all cities with improved scrolling."""
    print("\n" + "="*70)
    print(" STEP 1: Collecting URLs from Google Maps")
    print("="*70)
    print(f"🎯 Target: {target} listings")
    print(f"📍 State: {state}")
    print(f"🏙️ Cities: {len(cities)}")
    print(f"💼 Profession: {profession}")
    print(f"📜 Scrolling: Medium pace ({SCROLL_DELAY:.1f}s between scrolls)")
    print("="*70)
    
    all_urls = set()
    
    for idx, city in enumerate(cities, 1):
        print(f"\n{'─'*50}")
        print(f"City {idx}/{len(cities)}: {city}")
        
        # Get URLs with improved scrolling
        urls = await collect_urls_optimized(page, profession, state, city, limit=120)
        
        new_urls = [url for url in urls if url not in all_urls]
        all_urls.update(new_urls)
        
        print(f"   📈 New unique URLs from this city: {len(new_urls)}")
        print(f"   📊 Total unique URLs so far: {len(all_urls)}")
        
        # Check if we have enough
        if len(all_urls) >= target:
            print(f"\n✅ Collected enough URLs ({len(all_urls)} >= {target})")
            break
        
        # Wait between cities to avoid rate limiting
        if idx < len(cities) and len(all_urls) < target:
            delay = random.uniform(8, 12)
            print(f"   ⏳ Waiting {delay:.1f}s before next city...")
            await asyncio.sleep(delay)
    
    return list(all_urls)

# ============================================================================
# EXCEL SETUP
# ============================================================================

def setup_excel_fast(filename):
    """Create Excel workbook (minimal styling for speed)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Email Extracts"
    
    # Simple headers (faster)
    for col_idx, col_name in enumerate(COLUMNS, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    # Set column widths
    widths = [6, 35, 25, 10, 10, 45, 18, 50, 60, 50, 50, 35, 20, 15, 60, 20]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    wb.save(filename)
    return wb, ws

# ============================================================================
# BROWSER SETUP
# ============================================================================

async def setup_browser_optimized():
    """Setup Chrome browser with optimized settings."""
    print("🚀 Starting browser...")
    
    playwright = await async_playwright().start()
    
    browser = await playwright.chromium.launch(
        headless=False,
        args=[
            '--start-maximized',
            '--disable-blink-features=AutomationControlled',
            '--disable-dev-shm-usage',
            '--no-sandbox',
            '--disable-gpu',
            '--disable-accelerated-2d-canvas',
            '--disable-features=IsolateOrigins,site-per-process',
        ]
    )
    
    context = await browser.new_context(
        viewport={'width': random.choice([1366, 1440, 1920]), 
                  'height': random.choice([768, 900, 1080])},
        user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        locale='en-US',
        ignore_https_errors=True,
    )
    
    await context.add_init_script("""
        Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
        Object.defineProperty(navigator, 'plugins', {get: () => [1, 2, 3, 4, 5]});
        window.chrome = { runtime: {} };
    """)
    
    page = await context.new_page()
    
    print("   ✅ Browser ready")
    return playwright, browser, context, page

# ============================================================================
# MAIN FUNCTION
# ============================================================================

async def main():
    global saved_records_count
    
    print("\n" + "="*70)
    print("  OPTIMIZED BUT THOROUGH EMAIL SCRAPER")
    print("  ✓ Checks 6 pages per website (auto-discovery + fallback)")
    print("  ✓ ALWAYS checks Facebook")
    print("  ✓ IMPROVED SCROLLING - Gets ALL Google Maps results (up to 120)")
    print("  ✓ Optimized speed - won't skip data")
    print("="*70)
    
    # Get filename
    print("\n📁 FILENAME SETUP:")
    default_filename = f"emails_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    print(f"   Default: {default_filename}")
    custom_name = input("   Enter filename (or press Enter for default): ").strip()
    
    if custom_name:
        if not custom_name.endswith('.xlsx'):
            custom_name += '.xlsx'
        output_file = os.path.join(os.path.expanduser("~"), "Desktop", custom_name)
    else:
        output_file = os.path.join(os.path.expanduser("~"), "Desktop", default_filename)
    
    print(f"   💾 Output: {os.path.basename(output_file)}")
    
    # Get profession
    profession = input("\n💼 Profession (Doctor, Dentist, Lawyer): ").strip() or "Doctor"
    
    # Get state
    print("\n📌 States: CA, NY, TX, FL, IL")
    state_input = input("Enter state abbreviation: ").strip().upper()
    state_map = {"CA": "California", "NY": "New York", "TX": "Texas", "FL": "Florida", "IL": "Illinois"}
    state = state_map.get(state_input, "California")
    
    cities = US_STATES.get(state, ["Los Angeles"])
    
    # Get target
    try:
        target = int(input(f"\n📊 How many listings? (default 25): ").strip() or "25")
    except:
        target = 25
    
    print(f"\n🎯 Target: {target} listings with emails")
    print(f"📍 State: {state}")
    print(f"🏙️ Cities: {len(cities)}")
    print(f"📄 Website: {MAX_PAGES_TO_CHECK} pages per site (auto-discovery)")
    print(f"📘 Facebook: ALWAYS checked")
    print(f"📜 Google Maps: Medium-pace scrolling (gets ALL results up to 120 per city)")
    print(f"⏱️  Pace: {BETWEEN_LISTINGS:.1f}s between listings (optimized)")
    print(f"💾 Output: Desktop/{os.path.basename(output_file)}")
    
    playwright = None
    browser = None
    
    try:
        playwright, browser, context, page = await setup_browser_optimized()
        
        # Collect URLs with improved scrolling
        all_urls = await collect_all_urls_optimized(page, profession, state, cities, target * 2)
        
        if not all_urls:
            print("\n❌ No URLs found!")
            return
        
        random.shuffle(all_urls)
        
        # Setup Excel
        wb, ws = setup_excel_fast(output_file)
        saved_records_count = 0
        
        print("\n" + "="*70)
        print(" STEP 2: Extracting Emails (Optimized)")
        print("="*70)
        print(f"📋 URLs to check: {len(all_urls)}")
        print(f"📄 Website: {MAX_PAGES_TO_CHECK} pages each")
        print(f"📘 Facebook: ALWAYS checking")
        print(f"⏱️  {BETWEEN_LISTINGS:.1f}s between listings\n")
        
        for idx, url in enumerate(all_urls, 1):
            city = random.choice(cities)
            
            result = await scrape_listing_optimized(page, url, city, state, idx, len(all_urls), wb, ws, output_file)
            
            if result:
                print(f"   📊 Progress: {saved_records_count}/{target}")
                
                if saved_records_count >= target:
                    print(f"\n🎉 Target reached! Saved {saved_records_count} records.")
                    break
            
            if idx < len(all_urls) and saved_records_count < target:
                delay = random.uniform(BETWEEN_LISTINGS - 1, BETWEEN_LISTINGS + 1)
                print(f"   ⏳ Waiting {delay:.1f}s before next...")
                await asyncio.sleep(delay)
        
        # Final summary
        print("\n" + "="*70)
        print("  SCRAPING COMPLETED!")
        print("="*70)
        print(f"   ✅ Saved: {saved_records_count} records")
        print(f"   📁 File: {os.path.basename(output_file)}")
        print(f"   📂 Location: Desktop")
        print("="*70)
        
    except KeyboardInterrupt:
        print("\n\n🛑 Stopped by user")
        print(f"   💾 Saved {saved_records_count} records")
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        if browser:
            await browser.close()
        if playwright:
            await playwright.stop()
        print("\n✅ Done!")

if __name__ == "__main__":
    asyncio.run(main())











# this is updated version of upper
#!/usr/bin/env python3
# """
# COMPLETE EMAIL SCRAPER - CHECKS ALL WEBSITE PAGES
# - Checks ALL pages on website (unlimited crawling)
# - ALWAYS checks Facebook
# - Gets ALL Google Maps results (up to 120 per search)
# - Properly handles max_pages=None for unlimited discovery
# """

# import asyncio
# import re
# import os
# import random
# from datetime import datetime
# from typing import List, Dict, Optional, Set, Tuple
# from urllib.parse import urlparse, urljoin
# from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError
# import openpyxl
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# # ============================================================================
# # CONFIGURATION - OPTIMIZED BUT THOROUGH
# # ============================================================================

# # DON'T SKIP ANYTHING - Scrape all listings
# SKIP_LISTINGS_CONTAINING = []  # EMPTY - Scrape everything!

# # Only skip obvious junk emails (won't skip real emails)
# SKIP_EMAILS_CONTAINING = [
#     "sentry", "mailer-daemon", "postmaster", "noreply", "no-reply",
#     "donotreply", "example.com", "test.com", "placeholder"
# ]

# # Skip only obvious spam domains
# SKIP_EMAIL_DOMAINS = [
#     "wixpress.com", "sentry.io", "example.com", "example.org"
# ]

# # OPTIMIZED TIMING - Fast enough but won't get blocked
# BETWEEN_LISTINGS = random.uniform(5.0, 8.0)  # 5-8 seconds (optimized)
# BETWEEN_PAGES = random.uniform(1.0, 1.5)     # 1-1.5 seconds between pages
# NAVIGATION_TIMEOUT = 45000                   # 45 seconds (optimized)
# PAGE_LOAD_TIMEOUT = 20000                    # 20 seconds per page
# MAX_RETRIES = 2                              # Retry failed pages twice

# # PAGE DISCOVERY SETTINGS - CHECK ALL PAGES
# MAX_PAGES_TO_CHECK = None      # None = Check ALL discovered pages!
# CRAWL_DEPTH = 2                # How deep to crawl (1=homepage, 2=pages linked from homepage)
# MAX_TOTAL_PAGES = None         # None = No limit, or set number like 100 to limit

# # SCROLLING SETTINGS - Gets ALL results
# SCROLL_DELAY = random.uniform(2.0, 2.8)      # Medium pace scrolling (2-2.8 seconds)
# MAX_SCROLL_ATTEMPTS = 30                      # Enough for ~120 results
# NO_NEW_RESULTS_LIMIT = 4                      # Stop after 4 scrolls with no new results

# # Parallel settings (for concurrent page checks - FASTER but safe)
# MAX_CONCURRENT_PAGES = 2  # Check 2 pages at a time max (safe concurrency)

# # US States Database
# US_STATES = {
#     "California": ["Los Angeles", "San Diego", "San Francisco", "Sacramento"],
#     "New York": ["New York", "Buffalo", "Rochester"],
#     "Texas": ["Houston", "San Antonio", "Dallas", "Austin"],
#     "Florida": ["Miami", "Orlando", "Tampa", "Jacksonville"],
#     "Illinois": ["Chicago", "Aurora", "Rockford", "Joliet", "Naperville", "Springfield", "Peoria", "Elgin"],
# }

# # Google Maps Selectors
# SELECTORS = {
#     "results_panel": 'div[role="feed"]',
#     "result_link": 'a[href*="/maps/place/"]',
#     "name": 'h1.DUwDvf, h1[aria-level="1"]',
#     "rating": 'div.F7nice span[aria-hidden="true"]',
#     "reviews": 'button[aria-label*="reviews"] span',
#     "address": 'button[data-item-id="address"] div.Io6YTe',
#     "phone": 'button[data-item-id*="phone"] div.Io6YTe',
#     "website": 'a[data-item-id="authority"]',
#     "category": 'button[jsaction*="category"] span',
# }

# # Excel columns
# COLUMNS = [
#     "No.", "Listing Name", "Category", "Rating", "Reviews", "Address",
#     "Phone", "Website", "All Emails Found", "Website Emails",
#     "Facebook Emails", "Primary Email", "City", "State", "Google Maps URL", "Scrape Date"
# ]

# # FALLBACK PAGES - Used ONLY when auto-discovery fails
# FALLBACK_PAGES = [
#     "", "contact", "contact-us", "about", "about-us", "team", "our-team",
#     "meet-the-team", "staff", "doctors", "providers", "location", "hours"
# ]

# # ============================================================================
# # COMPLETE PAGE DISCOVERY - FINDS ALL PAGES (NO LIMIT)
# # ============================================================================

# async def discover_all_pages(page, base_url: str) -> List[str]:
#     """
#     DISCOVERS ALL PAGES on a website by crawling links.
#     No page limit - finds everything up to MAX_TOTAL_PAGES if set.
#     """
#     discovered_urls = set()
#     discovered_urls.add(base_url)
    
#     # Queue for BFS crawling
#     urls_to_visit = [base_url]
#     visited_urls = set()
#     current_depth = 0
    
#     print(f"   🔍 Crawling website to discover ALL pages...")
    
#     while urls_to_visit and current_depth < CRAWL_DEPTH:
#         current_batch = urls_to_visit.copy()
#         urls_to_visit = []
#         current_depth += 1
        
#         print(f"      Depth {current_depth}: Checking {len(current_batch)} pages...")
        
#         for page_url in current_batch[:30]:  # Limit per batch to avoid overload
#             if page_url in visited_urls:
#                 continue
                
#             visited_urls.add(page_url)
            
#             try:
#                 await page.goto(page_url, timeout=20000, wait_until="domcontentloaded")
#                 await asyncio.sleep(random.uniform(0.5, 1.0))
                
#                 # Find all links on this page
#                 new_links = await page.evaluate(f'''
#                     () => {{
#                         const links = new Set();
#                         const baseUrl = '{base_url}';
#                         const allLinks = document.querySelectorAll('a[href]');
#                         const skipExtensions = ['.jpg', '.jpeg', '.png', '.gif', '.pdf', '.css', '.js', '.xml', '.json', '.zip'];
                        
#                         for (let link of allLinks) {{
#                             let href = link.href;
#                             if (href && href.startsWith(baseUrl) && 
#                                 !href.includes('#') && 
#                                 !href.includes('?')) {{
                                
#                                 href = href.replace(/\\/$/, '');
#                                 const isContent = !skipExtensions.some(ext => href.toLowerCase().endsWith(ext));
#                                 if (isContent && href !== window.location.href) {{
#                                     links.add(href);
#                                 }}
#                             }}
#                         }}
#                         return Array.from(links);
#                     }}
#                 ''')
                
#                 for new_url in new_links:
#                     if new_url not in discovered_urls:
#                         discovered_urls.add(new_url)
#                         if current_depth < CRAWL_DEPTH:
#                             urls_to_visit.append(new_url)
                
#                 # Check if we've hit the max limit
#                 if MAX_TOTAL_PAGES and len(discovered_urls) >= MAX_TOTAL_PAGES:
#                     print(f"      Reached max page limit ({MAX_TOTAL_PAGES})")
#                     break
                
#             except Exception as e:
#                 continue
        
#         if MAX_TOTAL_PAGES and len(discovered_urls) >= MAX_TOTAL_PAGES:
#             break
    
#     # Convert to list
#     result = list(discovered_urls)
    
#     print(f"   ✅ Discovered {len(result)} total pages on this website")
    
#     # Show sample of discovered pages
#     for i, url in enumerate(result[:10], 1):
#         page_name = url.replace(base_url, '').strip('/')
#         if not page_name:
#             page_name = "homepage"
#         print(f"      {i}. {page_name[:60]}")
#     if len(result) > 10:
#         print(f"      ... and {len(result) - 10} more pages")
    
#     return result

# # ============================================================================
# # FALLBACK PAGE DISCOVERY (When crawling fails)
# # ============================================================================

# async def discover_fallback_pages(page, base_url: str, max_pages: int = 6) -> List[str]:
#     """
#     FALLBACK page discovery - used when crawling fails.
#     Uses common page names to find emails.
#     """
#     discovered_urls = set()
#     discovered_urls.add(base_url)
    
#     # Add fallback pages
#     for page_path in FALLBACK_PAGES[:max_pages - 1]:
#         if page_path:
#             full_url = urljoin(base_url, page_path)
#             if full_url not in discovered_urls:
#                 discovered_urls.add(full_url)
    
#     result = list(discovered_urls)[:max_pages]
#     print(f"   📄 Using fallback - checking {len(result)} common pages")
#     return result

# # ============================================================================
# # OPTIMIZED PAGE DISCOVERY (Handles None correctly)
# # ============================================================================

# async def discover_pages_optimized(page, base_url: str, max_pages: int = 6) -> List[str]:
#     """
#     OPTIMIZED page discovery:
#     - If max_pages = None: Discover ALL pages (unlimited)
#     - If max_pages is a number: Limit to that many pages
#     - Falls back to common pages if crawling fails
#     """
    
#     # CASE 1: Check ALL pages (unlimited)
#     if max_pages is None or max_pages == 0:
#         try:
#             return await discover_all_pages(page, base_url)
#         except Exception as e:
#             print(f"      ⚠️ Full crawl failed: {str(e)[:50]}")
#             print(f"      📋 Using fallback pages instead")
#             return await discover_fallback_pages(page, base_url, 15)
    
#     # CASE 2: Check limited number of pages
#     discovered_urls = set()
#     discovered_urls.add(base_url)
    
#     try:
#         # Fast homepage load
#         await page.goto(base_url, timeout=15000, wait_until="domcontentloaded")
#         await asyncio.sleep(random.uniform(0.5, 0.8))
        
#         # Quick link extraction
#         links = await page.evaluate('''
#             () => {
#                 const links = new Set();
#                 const baseUrl = window.location.origin;
#                 const allLinks = document.querySelectorAll('a[href]');
#                 const skipExtensions = ['.jpg', '.png', '.gif', '.pdf', '.css', '.js', '.xml'];
                
#                 for (let link of allLinks) {
#                     let href = link.href;
#                     if (href && href.startsWith(baseUrl) && 
#                         !href.includes('#') && 
#                         !href.includes('?') &&
#                         href !== window.location.href) {
                        
#                         href = href.replace(/\\/$/, '');
#                         const isContent = !skipExtensions.some(ext => href.toLowerCase().endsWith(ext));
#                         if (isContent) links.add(href);
#                     }
#                     if (links.size >= 20) break;
#                 }
#                 return Array.from(links);
#             }
#         ''')
        
#         # Add discovered links (up to max_pages)
#         for link in links[:max_pages - 1]:
#             if link and link != base_url:
#                 discovered_urls.add(link)
                
#         if len(discovered_urls) > 1:
#             print(f"      ✅ Found {len(discovered_urls) - 1} internal pages")
#         else:
#             print(f"      ⚠️ No internal links found, using fallback")
            
#     except Exception as e:
#         print(f"      ⚠️ Auto-discovery failed: {str(e)[:40]}")
    
#     # Add fallback pages if needed
#     if len(discovered_urls) < max_pages:
#         needed = max_pages - len(discovered_urls)
#         for page_path in FALLBACK_PAGES[:needed]:
#             if page_path:
#                 full_url = urljoin(base_url, page_path)
#                 if full_url not in discovered_urls:
#                     discovered_urls.add(full_url)
    
#     # Convert to list and return
#     result = list(discovered_urls)[:max_pages]
    
#     # Brief summary
#     print(f"      📄 Checking {len(result)} pages", end="")
#     if len(discovered_urls) > 1:
#         print(f" ({len(discovered_urls) - 1} discovered)")
#     else:
#         print(f" (using fallback)")
    
#     return result

# # ============================================================================
# # OPTIMIZED WEBSITE EMAIL EXTRACTION
# # ============================================================================

# async def extract_emails_from_website_optimized(page, website_url: str, listing_name: str) -> Tuple[List[str], List[str]]:
#     """
#     OPTIMIZED email extraction:
#     - Discovers pages (ALL if MAX_PAGES_TO_CHECK is None)
#     - Checks each page for emails
#     - No data skipped
#     """
#     if not website_url:
#         return [], []
    
#     base_url = website_url.rstrip('/')
#     email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
    
#     # Discover pages (handles None correctly for unlimited pages)
#     pages_to_check = await discover_pages_optimized(page, base_url, MAX_PAGES_TO_CHECK)
    
#     # Calculate total pages for display
#     total_pages = len(pages_to_check)
#     print(f"   🔍 Scanning {total_pages} pages for emails...")
    
#     # Check pages sequentially (safer, avoids detection)
#     all_emails = set()
#     successful_pages = 0
    
#     for i, page_url in enumerate(pages_to_check, 1):
#         try:
#             # Navigate to page
#             await page.goto(page_url, timeout=PAGE_LOAD_TIMEOUT, wait_until="domcontentloaded")
#             await asyncio.sleep(random.uniform(0.3, 0.5))
            
#             # Get content
#             content = await page.content()
#             emails = re.findall(email_pattern, content)
            
#             for email in emails:
#                 email_lower = email.lower()
                
#                 # Skip junk
#                 if any(x in email_lower for x in SKIP_EMAILS_CONTAINING):
#                     continue
                
#                 domain = email.split('@')[-1] if '@' in email else ""
#                 if any(skip in domain for skip in SKIP_EMAIL_DOMAINS):
#                     continue
                
#                 all_emails.add(email)
            
#             successful_pages += 1
            
#             # Progress indicator
#             if total_pages > 20:
#                 if i % 10 == 0 or i == total_pages:
#                     print(f"      Checked {i}/{total_pages} pages...", end="\r")
#             else:
#                 if i % 3 == 0:
#                     print(f"      Checked {i}/{total_pages} pages...", end="\r")
            
#         except Exception:
#             continue
    
#     print(f"      ✅ Checked {successful_pages}/{total_pages} pages successfully    ")
    
#     email_list = list(all_emails)
    
#     if email_list:
#         print(f"   📧 Found {len(email_list)} email(s) on website")
#         for e in email_list[:3]:
#             print(f"      - {e}")
#     else:
#         print(f"   📧 No emails found on website")
    
#     return email_list, email_list

# # ============================================================================
# # OPTIMIZED FACEBOOK CHECK (Always checked)
# # ============================================================================

# async def extract_emails_from_facebook_optimized(page, listing_name: str, website_url: str) -> List[str]:
#     """
#     OPTIMIZED Facebook check:
#     - ALWAYS checked regardless of website results
#     - Quick timeout (10 seconds)
#     """
#     emails_found = set()
#     email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
    
#     # Clean listing name
#     name_clean = re.sub(r'[^\w\s]', '', listing_name.lower())
#     name_clean = re.sub(r'\s+', '', name_clean)
    
#     # Extract domain
#     domain = ""
#     if website_url:
#         try:
#             parsed = urlparse(website_url)
#             domain = parsed.netloc.replace('www.', '').split('.')[0]
#         except:
#             pass
    
#     # Facebook URLs to try (prioritized)
#     facebook_urls = [
#         f"https://www.facebook.com/{name_clean}",
#         f"https://www.facebook.com/{domain}",
#     ]
    
#     print(f"   🔍 Checking Facebook...")
    
#     for fb_url in facebook_urls[:2]:
#         try:
#             await page.goto(fb_url, timeout=10000, wait_until="domcontentloaded")
#             await asyncio.sleep(random.uniform(0.5, 0.8))
            
#             content = await page.content()
#             emails = re.findall(email_pattern, content)
            
#             for email in emails:
#                 email_lower = email.lower()
#                 if not any(x in email_lower for x in SKIP_EMAILS_CONTAINING):
#                     if 'facebook.com' not in email_lower and 'fbcdn' not in email_lower:
#                         emails_found.add(email)
            
#         except Exception:
#             continue
    
#     email_list = list(emails_found)
    
#     if email_list:
#         print(f"   📧 Found {len(email_list)} email(s) on Facebook")
#         for e in email_list[:2]:
#             print(f"      - {e}")
#     else:
#         print(f"   📧 No emails found on Facebook")
    
#     return email_list

# # ============================================================================
# # COMPLETE EMAIL EXTRACTION
# # ============================================================================

# async def extract_all_emails_optimized(page, website_url: str, listing_name: str) -> Tuple[List[str], List[str], List[str]]:
#     """Extract all emails using optimized methods."""
    
#     all_emails_set = set()
    
#     # Extract from website (handles ALL pages if MAX_PAGES_TO_CHECK is None)
#     website_emails, _ = await extract_emails_from_website_optimized(page, website_url, listing_name)
#     all_emails_set.update(website_emails)
    
#     # Brief pause before Facebook
#     await asyncio.sleep(random.uniform(0.5, 0.8))
    
#     # Extract from Facebook (ALWAYS)
#     facebook_emails = await extract_emails_from_facebook_optimized(page, listing_name, website_url)
#     all_emails_set.update(facebook_emails)
    
#     all_emails_list = list(all_emails_set)
    
#     # Determine primary email (quick)
#     primary_email = ""
#     if all_emails_list:
#         for email in all_emails_list:
#             if '.com' in email and not any(x in email.lower() for x in ["info", "admin", "contact"]):
#                 primary_email = email
#                 break
#         if not primary_email:
#             primary_email = all_emails_list[0]
    
#     print(f"\n   📊 EMAIL SUMMARY: {len(all_emails_list)} total (web: {len(website_emails)}, FB: {len(facebook_emails)})")
#     if primary_email:
#         print(f"      Primary: {primary_email}")
    
#     return all_emails_list, website_emails, facebook_emails

# # ============================================================================
# # SCRAPE LISTING (Optimized)
# # ============================================================================

# async def scrape_listing_optimized(page, url: str, city: str, state: str, index: int, total: int,
#                                     wb, ws, filename: str, retry_count: int = 0) -> Optional[Dict]:
#     """Scrape a single listing with optimized settings."""
    
#     print(f"\n[{index}/{total}] 📍 {city[:15]} - {url.split('/')[-2][:35]}...")
    
#     try:
#         # Fast Google Maps load
#         await page.goto(url, timeout=NAVIGATION_TIMEOUT, wait_until="domcontentloaded")
#         await asyncio.sleep(random.uniform(1.0, 1.5))
        
#         # Extract basic info (quick)
#         listing_name = await get_text_fast(page, SELECTORS["name"])
#         if not listing_name:
#             print(f"   ❌ No name found")
#             return None
        
#         print(f"   🏢 {listing_name[:55]}")
        
#         # Extract other details
#         category = await get_text_fast(page, SELECTORS["category"])
#         rating = await get_text_fast(page, SELECTORS["rating"])
#         reviews_raw = await get_text_fast(page, SELECTORS["reviews"])
#         reviews = re.sub(r'[^\d]', '', reviews_raw) if reviews_raw else ""
#         address = await get_text_fast(page, SELECTORS["address"])
#         phone = await get_text_fast(page, SELECTORS["phone"])
        
#         # Extract website
#         website = await extract_website_fast(page)
        
#         if not website:
#             print(f"   ⏭️ No website found")
#             return None
        
#         print(f"   🌐 {website[:70]}...")
        
#         # Create new page for email extraction
#         email_page = await page.context.new_page()
        
#         # Extract ALL emails (optimized - handles ALL pages)
#         all_emails, website_emails, facebook_emails = await extract_all_emails_optimized(
#             email_page, website, listing_name
#         )
        
#         await email_page.close()
        
#         if not all_emails:
#             print(f"   ❌ No emails found")
#             return None
        
#         # Prepare data
#         data = {
#             "listing_name": listing_name,
#             "category": category,
#             "rating": rating,
#             "reviews": reviews,
#             "address": address,
#             "phone": phone,
#             "website": website,
#             "all_emails": ", ".join(all_emails),
#             "website_emails": ", ".join(website_emails),
#             "facebook_emails": ", ".join(facebook_emails),
#             "primary_email": all_emails[0] if all_emails else "",
#             "city": city,
#             "state": state,
#             "url": url,
#             "scrape_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
#         }
        
#         print(f"\n   ✅ FOUND {len(all_emails)} emails! (Saved)")
        
#         # Save to Excel
#         save_to_excel_fast(wb, ws, data, filename)
        
#         return data
        
#     except Exception as e:
#         print(f"   ❌ Error: {str(e)[:60]}")
#         if retry_count < MAX_RETRIES:
#             print(f"   🔄 Retry {retry_count + 1}/{MAX_RETRIES}...")
#             await asyncio.sleep(3)
#             return await scrape_listing_optimized(page, url, city, state, index, total, wb, ws, filename, retry_count + 1)
#         return None

# # ============================================================================
# # FAST HELPER FUNCTIONS (Optimized)
# # ============================================================================

# async def extract_website_fast(page) -> str:
#     """Extract website URL quickly."""
#     selectors = ['a[data-item-id="authority"]', 'a[aria-label*="Website"]']
    
#     for selector in selectors:
#         try:
#             element = await page.query_selector(selector)
#             if element:
#                 href = await element.get_attribute("href")
#                 if href and ("http" in href or "www" in href):
#                     if href.startswith("//"):
#                         href = "https:" + href
#                     return href.rstrip('/')
#         except:
#             continue
#     return ""

# async def get_text_fast(page, *selectors) -> str:
#     """Get text quickly with fallback."""
#     for selector in selectors:
#         try:
#             element = await page.query_selector(selector)
#             if element:
#                 text = await element.inner_text()
#                 if text and text.strip():
#                     return text.strip()[:500]
#         except:
#             continue
#     return ""

# saved_records_count = 0

# def save_to_excel_fast(wb, ws, data: Dict, filename: str):
#     """Save data to Excel quickly."""
#     global saved_records_count
    
#     next_row = ws.max_row + 1
    
#     values = [
#         next_row - 1,
#         data.get("listing_name", ""),
#         data.get("category", ""),
#         data.get("rating", ""),
#         data.get("reviews", ""),
#         data.get("address", ""),
#         data.get("phone", ""),
#         data.get("website", ""),
#         data.get("all_emails", ""),
#         data.get("website_emails", ""),
#         data.get("facebook_emails", ""),
#         data.get("primary_email", ""),
#         data.get("city", ""),
#         data.get("state", ""),
#         data.get("url", ""),
#         data.get("scrape_date", ""),
#     ]
    
#     for col_idx, value in enumerate(values, 1):
#         ws.cell(row=next_row, column=col_idx, value=value)
    
#     try:
#         wb.save(filename)
#         saved_records_count += 1
#     except Exception as e:
#         print(f"   ⚠️ Save error: {e}")

# # ============================================================================
# # IMPROVED URL COLLECTION WITH PROPER SCROLLING - GETS ALL RESULTS
# # ============================================================================

# async def collect_urls_optimized(page, profession: str, state: str, city: str, limit: int = 120) -> List[str]:
#     """
#     IMPROVED URL collection with MEDIUM-PACED scrolling to get ALL results.
#     Google Maps shows ~120 results maximum per search.
#     """
#     print(f"\n   🌆 Searching: {profession} in {city}, {state}")
    
#     query = f"{profession} in {city}, {state}"
#     search_url = f"https://www.google.com/maps/search/{query.replace(' ', '+')}/"
    
#     try:
#         await page.goto(search_url, timeout=30000, wait_until="domcontentloaded")
#         await asyncio.sleep(random.uniform(2.5, 3.5))
#     except:
#         print(f"   ⚠️ Failed to load search")
#         return []
    
#     # Handle cookie consent if present
#     try:
#         accept_btn = await page.query_selector('button:has-text("Accept all")')
#         if accept_btn:
#             await accept_btn.click()
#             await asyncio.sleep(1)
#     except:
#         pass
    
#     # Get the results panel
#     results_panel = await page.query_selector('div[role="feed"]')
#     if not results_panel:
#         print(f"   ⚠️ Could not find results panel, trying page scroll...")
#         results_panel = None
    
#     urls = set()
#     previous_count = 0
#     no_new_count = 0
#     scroll_attempts = 0
    
#     print(f"      📜 Starting medium-pace scrolling to collect ALL results...")
#     print(f"      ⏱️  {SCROLL_DELAY:.1f}s between scrolls")
    
#     while len(urls) < limit and scroll_attempts < MAX_SCROLL_ATTEMPTS and no_new_count < NO_NEW_RESULTS_LIMIT:
        
#         # Extract current URLs
#         links = await page.query_selector_all('a[href*="/maps/place/"]')
        
#         for link in links:
#             href = await link.get_attribute("href")
#             if href and "/maps/place/" in href:
#                 clean_url = href.split("?")[0]
#                 urls.add(clean_url)
        
#         # Show progress
#         print(f"      📊 Collected {len(urls)} URLs...", end="\r")
        
#         # Check if we got new results
#         if len(urls) == previous_count:
#             no_new_count += 1
#             if no_new_count == 1:
#                 print(f"\n      ⏸️ No new results after scroll {scroll_attempts + 1}")
#         else:
#             new_found = len(urls) - previous_count
#             print(f"\n      ✅ +{new_found} new results (Total: {len(urls)})")
#             no_new_count = 0
        
#         previous_count = len(urls)
        
#         # Stop if we have enough or reached end
#         if len(urls) >= limit:
#             break
        
#         if no_new_count >= NO_NEW_RESULTS_LIMIT:
#             print(f"\n      🏁 Reached end of results (no new listings after {NO_NEW_RESULTS_LIMIT} scrolls)")
#             break
        
#         # SCROLLING - Medium pace
#         if results_panel:
#             try:
#                 await results_panel.evaluate("element => element.scrollTop = element.scrollHeight")
#             except:
#                 await page.evaluate("window.scrollBy(0, 800)")
#         else:
#             await page.evaluate("window.scrollBy(0, 800)")
        
#         # Medium pace delay
#         await asyncio.sleep(SCROLL_DELAY)
#         scroll_attempts += 1
        
#         # Progress update every 5 scrolls
#         if scroll_attempts % 5 == 0:
#             print(f"      📜 Scroll {scroll_attempts}/{MAX_SCROLL_ATTEMPTS} - {len(urls)} URLs so far")
    
#     print(f"\n      ✅ Scrolling complete! Final count: {len(urls)} URLs")
    
#     # Show statistics
#     if len(urls) >= 100:
#         print(f"      🎯 Excellent! Got {len(urls)} results (near maximum of ~120)")
#     elif len(urls) >= 60:
#         print(f"      👍 Good! Got {len(urls)} results")
#     elif len(urls) >= 30:
#         print(f"      📊 Got {len(urls)} results - moderate coverage")
#     else:
#         print(f"      ⚠️ Got only {len(urls)} results - try a broader search term")
    
#     return list(urls)

# # ============================================================================
# # URL COLLECTION FROM ALL CITIES
# # ============================================================================

# async def collect_all_urls_optimized(page, profession: str, state: str, cities: List[str], target: int) -> List[str]:
#     """Collect URLs from all cities with improved scrolling."""
#     print("\n" + "="*70)
#     print(" STEP 1: Collecting URLs from Google Maps")
#     print("="*70)
#     print(f"🎯 Target: {target} listings")
#     print(f"📍 State: {state}")
#     print(f"🏙️ Cities: {len(cities)}")
#     print(f"💼 Profession: {profession}")
#     print(f"📜 Scrolling: Medium pace ({SCROLL_DELAY:.1f}s between scrolls)")
#     print("="*70)
    
#     all_urls = set()
    
#     for idx, city in enumerate(cities, 1):
#         print(f"\n{'─'*50}")
#         print(f"City {idx}/{len(cities)}: {city}")
        
#         # Get URLs with improved scrolling
#         urls = await collect_urls_optimized(page, profession, state, city, limit=120)
        
#         new_urls = [url for url in urls if url not in all_urls]
#         all_urls.update(new_urls)
        
#         print(f"   📈 New unique URLs from this city: {len(new_urls)}")
#         print(f"   📊 Total unique URLs so far: {len(all_urls)}")
        
#         # Check if we have enough
#         if len(all_urls) >= target:
#             print(f"\n✅ Collected enough URLs ({len(all_urls)} >= {target})")
#             break
        
#         # Wait between cities to avoid rate limiting
#         if idx < len(cities) and len(all_urls) < target:
#             delay = random.uniform(8, 12)
#             print(f"   ⏳ Waiting {delay:.1f}s before next city...")
#             await asyncio.sleep(delay)
    
#     return list(all_urls)

# # ============================================================================
# # EXCEL SETUP
# # ============================================================================

# def setup_excel_fast(filename):
#     """Create Excel workbook (minimal styling for speed)."""
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Email Extracts"
    
#     # Simple headers (faster)
#     for col_idx, col_name in enumerate(COLUMNS, 1):
#         ws.cell(row=1, column=col_idx, value=col_name)
    
#     # Set column widths
#     widths = [6, 35, 25, 10, 10, 45, 18, 50, 60, 50, 50, 35, 20, 15, 60, 20]
#     for i, width in enumerate(widths, 1):
#         ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
#     wb.save(filename)
#     return wb, ws

# # ============================================================================
# # BROWSER SETUP
# # ============================================================================

# async def setup_browser_optimized():
#     """Setup Chrome browser with optimized settings."""
#     print("🚀 Starting browser...")
    
#     playwright = await async_playwright().start()
    
#     browser = await playwright.chromium.launch(
#         headless=False,
#         args=[
#             '--start-maximized',
#             '--disable-blink-features=AutomationControlled',
#             '--disable-dev-shm-usage',
#             '--no-sandbox',
#             '--disable-gpu',
#             '--disable-accelerated-2d-canvas',
#             '--disable-features=IsolateOrigins,site-per-process',
#         ]
#     )
    
#     context = await browser.new_context(
#         viewport={'width': random.choice([1366, 1440, 1920]), 
#                   'height': random.choice([768, 900, 1080])},
#         user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
#         locale='en-US',
#         ignore_https_errors=True,
#     )
    
#     await context.add_init_script("""
#         Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
#         Object.defineProperty(navigator, 'plugins', {get: () => [1, 2, 3, 4, 5]});
#         window.chrome = { runtime: {} };
#     """)
    
#     page = await context.new_page()
    
#     print("   ✅ Browser ready")
#     return playwright, browser, context, page

# # ============================================================================
# # MAIN FUNCTION
# # ============================================================================

# async def main():
#     global saved_records_count
    
#     print("\n" + "="*70)
#     print("  COMPLETE EMAIL SCRAPER")
#     print(f"  ✓ Website pages: {'ALL pages (unlimited)' if MAX_PAGES_TO_CHECK is None else f'Up to {MAX_PAGES_TO_CHECK} pages'}")
#     print("  ✓ ALWAYS checks Facebook")
#     print("  ✓ Gets ALL Google Maps results (up to 120 per city)")
#     print("="*70)
    
#     # Get filename
#     print("\n📁 FILENAME SETUP:")
#     default_filename = f"emails_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
#     print(f"   Default: {default_filename}")
#     custom_name = input("   Enter filename (or press Enter for default): ").strip()
    
#     if custom_name:
#         if not custom_name.endswith('.xlsx'):
#             custom_name += '.xlsx'
#         output_file = os.path.join(os.path.expanduser("~"), "Desktop", custom_name)
#     else:
#         output_file = os.path.join(os.path.expanduser("~"), "Desktop", default_filename)
    
#     print(f"   💾 Output: {os.path.basename(output_file)}")
    
#     # Get profession
#     profession = input("\n💼 Profession (Doctor, Dentist, Lawyer): ").strip() or "Doctor"
    
#     # Get state
#     print("\n📌 States: CA, NY, TX, FL, IL")
#     state_input = input("Enter state abbreviation: ").strip().upper()
#     state_map = {"CA": "California", "NY": "New York", "TX": "Texas", "FL": "Florida", "IL": "Illinois"}
#     state = state_map.get(state_input, "California")
    
#     cities = US_STATES.get(state, ["Los Angeles"])
    
#     # Get target
#     try:
#         target = int(input(f"\n📊 How many listings? (default 25): ").strip() or "25")
#     except:
#         target = 25
    
#     print(f"\n🎯 Target: {target} listings with emails")
#     print(f"📍 State: {state}")
#     print(f"🏙️ Cities: {len(cities)}")
#     print(f"📄 Website: {'ALL pages (unlimited crawling)' if MAX_PAGES_TO_CHECK is None else f'{MAX_PAGES_TO_CHECK} pages per site'}")
#     print(f"📘 Facebook: ALWAYS checked")
#     print(f"📜 Google Maps: Medium-pace scrolling (gets ALL results up to 120 per city)")
#     print(f"⏱️  Pace: {BETWEEN_LISTINGS:.1f}s between listings")
#     print(f"💾 Output: Desktop/{os.path.basename(output_file)}")
    
#     playwright = None
#     browser = None
    
#     try:
#         playwright, browser, context, page = await setup_browser_optimized()
        
#         # Collect URLs with improved scrolling
#         all_urls = await collect_all_urls_optimized(page, profession, state, cities, target * 2)
        
#         if not all_urls:
#             print("\n❌ No URLs found!")
#             return
        
#         random.shuffle(all_urls)
        
#         # Setup Excel
#         wb, ws = setup_excel_fast(output_file)
#         saved_records_count = 0
        
#         print("\n" + "="*70)
#         print(" STEP 2: Extracting Emails")
#         print("="*70)
#         print(f"📋 URLs to check: {len(all_urls)}")
#         print(f"📄 Website: {'Scanning ALL discovered pages' if MAX_PAGES_TO_CHECK is None else f'Checking {MAX_PAGES_TO_CHECK} pages each'}")
#         print(f"📘 Facebook: ALWAYS checking")
#         print(f"⏱️  {BETWEEN_LISTINGS:.1f}s between listings\n")
        
#         for idx, url in enumerate(all_urls, 1):
#             city = random.choice(cities)
            
#             result = await scrape_listing_optimized(page, url, city, state, idx, len(all_urls), wb, ws, output_file)
            
#             if result:
#                 print(f"   📊 Progress: {saved_records_count}/{target}")
                
#                 if saved_records_count >= target:
#                     print(f"\n🎉 Target reached! Saved {saved_records_count} records.")
#                     break
            
#             if idx < len(all_urls) and saved_records_count < target:
#                 delay = random.uniform(BETWEEN_LISTINGS - 1, BETWEEN_LISTINGS + 1)
#                 print(f"   ⏳ Waiting {delay:.1f}s before next...")
#                 await asyncio.sleep(delay)
        
#         # Final summary
#         print("\n" + "="*70)
#         print("  SCRAPING COMPLETED!")
#         print("="*70)
#         print(f"   ✅ Saved: {saved_records_count} records")
#         print(f"   📁 File: {os.path.basename(output_file)}")
#         print(f"   📂 Location: Desktop")
#         print("="*70)
        
#     except KeyboardInterrupt:
#         print("\n\n🛑 Stopped by user")
#         print(f"   💾 Saved {saved_records_count} records")
#     except Exception as e:
#         print(f"\n❌ Error: {e}")
#         import traceback
#         traceback.print_exc()
#     finally:
#         if browser:
#             await browser.close()
#         if playwright:
#             await playwright.stop()
#         print("\n✅ Done!")

# if __name__ == "__main__":
#     asyncio.run(main())
